import os
import pandas as pd
import openpyxl
from lesson_functions import get_default_lesson_datas, empty_exam_result_row, MAX_TURKCE_QUESTIONS, MAX_MATEMATIK_QUESTIONS, MAX_FEN_QUESTIONS, MAX_INKILAP_QUESTIONS, MAX_DIN_QUESTIONS, MAX_INGILIZCE_QUESTIONS, TOTAL_QUESTIONS

# EXCEL FILE PATHS
current_script_directory = os.path.dirname(os.path.abspath(__file__))
SCHEDULE_EXCEL_PATH = os.path.join(current_script_directory, "weekly_schedule.xlsx")
EXAM_RESULT_EXCEL_PATH = os.path.join(current_script_directory, "exam_results.xlsx")
TEST_BOOK_LIST_PATH = os.path.join(current_script_directory, "test_book_list.xlsx")

# ----- connect_ai.py -----

# Get exam results from excel file and convert to string to send to AI model
def get_personal_data():
    if os.path.exists(EXAM_RESULT_EXCEL_PATH):
        exam_data = pd.read_excel(EXAM_RESULT_EXCEL_PATH)
        return exam_data.to_string()
    else:
        print("Personal data does not exist!")
        return "Personal data does not exist!"

def get_test_book_list():
    if os.path.exists(TEST_BOOK_LIST_PATH):
        book_data = pd.read_excel(TEST_BOOK_LIST_PATH)
        return book_data.to_string()
    else:
        print("Book list does not exist!")
        return "Book list does not exist!"

# Create a weekly schedule based on the AI's response -> text to excel file
def create_schedule(response:str):
    words = response.split()

    wb = openpyxl.Workbook()
    excelFile = wb.active
    excelFile.title = "Haftalık Program"

    weekDays = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
    hoursDic = {
                    "08:00": 2, "09:00": 3, "10:00": 4, 
                    "11:00": 5, "12:00": 6, "13:00": 7, 
                    "14:00": 8, "15:00": 9, "16:00": 10, 
                    "17:00": 11, "18:00": 12, "19:00": 13, 
                    "20:00": 14, "21:00": 15
                }

    # First column for hours
    excelFile.cell(row=1, column=1, value="Saatler") 
    for i in range(0, len(hoursDic)):
        excelFile.cell(row=i + 2, column=1, value=list(hoursDic.keys())[i])
        
    hour = None
    column = 1
    row = 2
    for word in words:
        if word in weekDays:
            column += 1
            row = 2  
            excelFile.cell(row=1, column=column, value=word)  
        else:
            if not (word in hoursDic):
                if word == "NULL":
                    excelFile.cell(row=hoursDic[hour], column=column, value="Boş Zaman")
                else:
                    excelFile.cell(row=hoursDic[hour], column=column, value=word)
                row += 1
            else:
                hour = word
            
    wb.save("weekly_schedule.xlsx")
    return "Ders programı oluşturuldu!"

# ----- app.py -----

# Read exam results excel file for main page
def read_exam_result_excel_file():
    lesson_labels = [] # lesson names
    success_percentages = []

    if os.path.exists(EXAM_RESULT_EXCEL_PATH):
        try:
            workbook_exam = openpyxl.load_workbook(EXAM_RESULT_EXCEL_PATH)
            sheet_exam = workbook_exam.active

            # Check and load column titles
            headers = [cell.value for cell in sheet_exam[1]]
            expected_headers = ["Deneme İsmi", "Türkçe Neti", "Matematik Neti", "Fen Neti", "İnkılap Neti", "Din Kültürü Neti", "İngilizce Neti", "Toplam Net"]

            if headers != expected_headers:
                print(f"Warning: {EXAM_RESULT_EXCEL_PATH} column's names are not like expected. Should be: {expected_headers}")
                (lesson_labels, success_percentages) = get_default_lesson_datas()
                
            else:
                total_turkce = .0
                total_matematik = .0
                total_fen = .0
                total_inkilap = .0
                total_din = .0
                total_ingilizce = .0
                total_toplam = .0
                exam_count = .0

                for row in sheet_exam.iter_rows(min_row=2, values_only=True): #first row is header

                    if len(row) >= 6 and row[0] is not None:
                        try:                            
                            total_turkce += (row[1] if row[1] is not None else 0)
                            total_matematik += (row[2] if row[2] is not None else 0)
                            total_fen += (row[3] if row[3] is not None else 0)
                            total_inkilap += (row[4] if row[4] is not None else 0)
                            total_din += (row[5] if row[5] is not None else 0)
                            total_ingilizce += (row[6] if row[6] is not None else 0)
                            total_toplam += (row[7] if row[7] is not None else 0)
                            exam_count += 1
                        except TypeError as e:
                            print(f"ERROR: {EXAM_RESULT_EXCEL_PATH} non-numeric value! Row: {row}, error: {e}")
                            continue

                if exam_count > 0:
                    avg_turkce = total_turkce / exam_count
                    avg_matematik = total_matematik / exam_count
                    avg_fen = total_fen / exam_count
                    avg_inkilap = total_inkilap / exam_count
                    avg_din = total_din / exam_count
                    avg_ingilizce = total_ingilizce / exam_count
                    avg_toplam = total_toplam / exam_count
                    
                    # calculate success percentages
                    # order is important: Türkçe, Matematik, Fen, Sosyal, Toplam
                    lesson_labels = ["Türkçe", "Matematik", "Fen", "İnkılap", "Din Kültürü", "İngilizce", "Toplam"]
                    success_percentages.append(round((avg_turkce / MAX_TURKCE_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_matematik / MAX_MATEMATIK_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_fen / MAX_FEN_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_inkilap / MAX_INKILAP_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_din / MAX_DIN_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_ingilizce / MAX_INGILIZCE_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_toplam / TOTAL_QUESTIONS) * 100, 2))
                else:
                    (lesson_labels, success_percentages) = get_default_lesson_datas()
            
        except Exception as e:
            print(f"ERROR: {e}")
            (lesson_labels, success_percentages) = get_default_lesson_datas()
    else:
        print(f"{EXAM_RESULT_EXCEL_PATH} does not exist!")
        (lesson_labels, success_percentages) = get_default_lesson_datas()
        
    return (lesson_labels, success_percentages)

# Read weekly schedule excel file for main page   
def read_weekly_schedule_excel_file():
    table_headers = []
    table_rows = []

    if os.path.exists(SCHEDULE_EXCEL_PATH):
        try:
            workbook_table = openpyxl.load_workbook(SCHEDULE_EXCEL_PATH)
            sheet_table = workbook_table.active

            table_headers = [cell.value for cell in sheet_table[1]]

            for row in sheet_table.iter_rows(min_row=2, values_only=True):
                # if value is None, replace with empty string, else convert to string
                cleaned_row = ["" if val is None else str(val) for val in row] 
                table_rows.append(cleaned_row)
            
        except Exception as e:
            print(f"It could not read table datas: {e}")
            table_headers = ["Error"]
            table_rows = [["Invalid Data"]]
    else:
        print(f"Not found excel file: {SCHEDULE_EXCEL_PATH}")
        table_headers = ["Haftalık program bulunamadı!"]
        table_rows = [["Notlarınızı girerek yapay zekaya program hazırlatabilirsiniz."]]
    
    return (table_headers, table_rows)

# Get exam results from excel file for add_score page
def get_exam_results():
    
    deneme_labels = []
    turkce_score = []
    matematik_score = []
    fen_score = []
    inkilap_score = []
    din_score = []
    ingilizce_score = []
    total_score = []
    
    if os.path.exists(EXAM_RESULT_EXCEL_PATH):
        try:
            workbook_exam = openpyxl.load_workbook(EXAM_RESULT_EXCEL_PATH)
            sheet_exam = workbook_exam.active
            for row in sheet_exam.iter_rows(min_row=2, values_only=True):
                
                if len(row) >= 6: # at least 6 columns
                    
                    if row[0] is None:
                        deneme_labels.append("Deneme İsmi Yok") # exam name is None
                    else:
                        deneme_labels.append(row[0]) # exam name

                    turkce_score.append(row[1] if row[1] is not None else 0)
                    matematik_score.append(row[2] if row[2] is not None else 0)
                    fen_score.append(row[3] if row[3] is not None else 0)
                    inkilap_score.append(row[4] if row[4] is not None else 0)
                    din_score.append(row[5] if row[5] is not None else 0)
                    ingilizce_score.append(row[6] if row[6] is not None else 0)
                    total_score.append(row[7] if row[7] is not None else 0)
                
        except Exception as e:
            print(f"Error: {e}")
            (deneme_labels, turkce_score, matematik_score, fen_score, inkilap_score, din_score, ingilizce_score, total_score) = empty_exam_result_row()
    else:
        print(f"Çizgi Grafik (deneme_neti.xlsx) Excel dosyası bulunamadı: {EXAM_RESULT_EXCEL_PATH}")
        (deneme_labels, turkce_score, matematik_score, fen_score, inkilap_score, din_score, ingilizce_score, total_score) = empty_exam_result_row()
        
    return (deneme_labels, turkce_score, matematik_score, fen_score, inkilap_score, din_score, ingilizce_score, total_score)

# Add new exam result to excel file for add_score page
def add_new_exam_result(exam_name, turkce_score, matematik_score, fen_score, inkilap_score, din_score, ingilizce_score, total_score) -> None:
    # If excel file does not exist, create it
    if not os.path.exists(EXAM_RESULT_EXCEL_PATH):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Ders Netleri"
        sheet.append(["Deneme İsmi", "Türkçe Neti", "Matematik Neti", "Fen Neti", "İnkılap Neti", "Din Kültürü Neti", "İngilizce Neti", "Toplam Net"])
        workbook.save(EXAM_RESULT_EXCEL_PATH)
    else:
        workbook = openpyxl.load_workbook(EXAM_RESULT_EXCEL_PATH)
        sheet = workbook.active

    # check if the exam name already exists
    found_row_idx = -1 # -1 means not found -> it is a FLAG
    for r_idx in range(2, sheet.max_row + 1): # 2. satırdan başla
        if sheet.cell(row=r_idx, column=1).value == exam_name: # find exam name
            found_row_idx = r_idx
            break
    
    if found_row_idx != -1: # exam name found
        sheet.cell(row=found_row_idx, column=2).value = turkce_score
        sheet.cell(row=found_row_idx, column=3).value = matematik_score
        sheet.cell(row=found_row_idx, column=4).value = fen_score
        sheet.cell(row=found_row_idx, column=5).value = inkilap_score
        sheet.cell(row=found_row_idx, column=6).value = din_score
        sheet.cell(row=found_row_idx, column=7).value = ingilizce_score
        sheet.cell(row=found_row_idx, column=8).value = total_score
    else: # exam name not found, append a new row
        new_row_data = [exam_name, turkce_score, matematik_score, fen_score, inkilap_score, din_score, ingilizce_score, total_score]
        sheet.append(new_row_data) # on last row
    
    workbook.save(EXAM_RESULT_EXCEL_PATH)
    