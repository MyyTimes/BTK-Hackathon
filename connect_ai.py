from dotenv import load_dotenv
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage
import pandas as pd
from openpyxl import Workbook
from app import get_max_question_number

START_HOUR = "08:00"
END_HOUR = "21:00"

# get exam results from excel file and convert to string to send to AI model
def get_personal_data(): 
    examData = pd.read_excel("exam_results.xlsx")
    return examData.to_string()

# create a weekly schedule based on the AI's response -> text to excel file
def create_schedule(response:str):
    words = response.split()

    wb = Workbook()
    excelFile = wb.active
    excelFile.title = "Haftalık Program"

    weekDays = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
    hoursDic = {
                    "08:00": 2, "09:00": 3, "10:00": 4, 
                    "11:00": 5, "12:00": 6, "13:00": 7, 
                    "14:00": 8, "15:00": 9, "16:00": 10, 
                    "17:00": 11, "18:00": 12, "19:00": 13, 
                    "20:00": 14, "21:00": 15
                }

    # First column for hours
    excelFile.cell(row=1, column=1, value="Saatler") 
    for i in range(0, len(hoursDic)):
        excelFile.cell(row=i + 2, column=1, value=list(hoursDic.keys())[i])
        
    hour = None
    column = 1
    row = 2
    for word in words:
        if word in weekDays:
            column += 1
            row = 2  
            excelFile.cell(row=1, column=column, value=word)  
        else:
            if not (word in hoursDic):
                if word == "NULL":
                    excelFile.cell(row=hoursDic[hour], column=column, value="Boş Zaman")
                else:
                    excelFile.cell(row=hoursDic[hour], column=column, value=word)
                row += 1
            else:
                hour = word
            
    wb.save("weekly_schedule.xlsx")
    
def give_answer(model, messages):
    response = model.invoke(messages)
    return response.content

def ai_answer(user_input):
    load_dotenv()

    model = ChatGoogleGenerativeAI(model="gemini-2.0-flash")
    messages = [
        SystemMessage(content=f"""You are an exam assistant of a student.
                                There are 4 lessons: Matematik, Turkce, Fen, Sosyal.
                                Question numbers are : {get_max_question_number()}.
                                Last exams results are : {get_personal_data()}.
                                You'll review the user's message. 
                                You should be realistic.
                                You can say negative feedbacks of the student if the scores are unsufficient.
                                Give some advice.
                                Write maximum 5 sentences."""),
        
        HumanMessage(content=user_input),
    ]

    return give_answer(model, messages)

# Create a weekly schedule based on the AI's response
def give_welcome_answer(model, messages):
    
    response = model.invoke(messages).content
    
    create_schedule(response)

    return response
    

# this function is for the welcome message of the AI assistant
# it will create a weekly schedule for the user based on their exam results
def ai_welcome_message():
    load_dotenv()

    model = ChatGoogleGenerativeAI(model="gemini-2.0-flash")
    messages = [
        SystemMessage(content=f"""You are an exam assistant helping a student create a personalized weekly study schedule based on their latest exam results.

                                    There are 4 lessons: Matematik, Türkçe, Fen, and Sosyal.

                                    The total number of questions per lesson is given by: {get_max_question_number()}.

                                    The student's recent exam results are: {get_personal_data()}.

                                    Create a detailed weekly study schedule from Monday (Pazartesi) to Sunday (Pazar).

                                    - Each day starts at {START_HOUR} and ends at {END_HOUR}.
                                    - Each hour should be represented in the schedule in 24-hour format with leading zeros, e.g. "08:00".
                                    - For each hour, first write the day of the week (in Turkish), followed by the lesson to study at that hour.
                                    - If no lesson is scheduled at a specific hour, write "NULL".
                                    - The output for each day should be a single line starting with the day name, followed by pairs of "hour lesson" separated by spaces.
                                    - Example output for one day:
                                    Pazartesi 08:00 Türkçe 09:00 Matematik 10:00 NULL 11:00 Fen 12:00 NULL 13:00 Sosyal 14:00 Matematik 15:00 NULL 16:00 Türkçe 17:00 NULL 18:00 Fen 19:00 NULL 20:00 Sosyal 21:00 NULL

                                    Only output the weekly schedule exactly in this format, with no extra text, explanation, or punctuation. This output will be parsed automatically, so please keep the format strict and consistent.

                                    """),

        
        HumanMessage(content="Bana haftalık bir çalışma programı hazırla. Saat 10.00 ile 16.00 arası okuldayım. O aralığa okul yaz."),
    ]

    return give_welcome_answer(model, messages)


