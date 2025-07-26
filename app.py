from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import openpyxl
import os
import connect_ai

app = Flask(__name__)
app.secret_key = "secret_key_A1W2S3E4D5F6" # for flash messages

# EXCEL
SCHEDULE_EXCEL_PATH = os.path.join(app.root_path, "weekly_schedule.xlsx")
EXAM_RESULT_EXCEL_PATH = os.path.join(app.root_path, "exam_results.xlsx")

# MAXIMUM QUESTIONS
MAX_TURKCE_QUESTIONS = 40
MAX_MATEMATIK_QUESTIONS = 40 
MAX_FEN_QUESTIONS = 20
MAX_SOCIAL_QUESTIONS = 20
TOTAL_QUESTIONS = 120

def GetMaxQuestionNumber():
    return {"Matematik": MAX_MATEMATIK_QUESTIONS, "Turkce" : MAX_TURKCE_QUESTIONS, "Fen" : MAX_FEN_QUESTIONS, "Sosyal" : MAX_SOCIAL_QUESTIONS}
def GetDefaultLessonDatas():
    lesson_labels = ["Türkçe", "Matematik", "Fen", "Sosyal", "Toplam"]
    success_percentages = [0, 0, 0, 0, 0]
    return lesson_labels, success_percentages

def EmptyExamResultRow():
    deneme_labels = ["Error"]
    turkce_score = [0]
    matematik_score = [0]
    fen_score = [0]
    sosyal_score = [0]
    total_score = [0] 
    return (deneme_labels, turkce_score, matematik_score, fen_score, sosyal_score, total_score)

# ---------- MAIN ROUTER ---------- 
@app.route('/')
def index(): # ---------- MAIN PAGE ----------
    
    # Read exam results excel file
    lesson_labels = [] # lesson names
    success_percentages = []

    if os.path.exists(EXAM_RESULT_EXCEL_PATH):
        try:
            workbook_exam = openpyxl.load_workbook(EXAM_RESULT_EXCEL_PATH)
            sheet_exam = workbook_exam.active

            # Check and load column titles
            headers = [cell.value for cell in sheet_exam[1]]
            expected_headers = ["Deneme İsmi", "Türkçe Neti", "Matematik Neti", "Fen Neti", "Sosyal Neti", "Toplam Net"]

            if headers != expected_headers:
                print(f"Warning: {EXAM_RESULT_EXCEL_PATH} column's names are not like expected. Should be: {expected_headers}")
                (lesson_labels, success_percentages) = GetDefaultLessonDatas()
                
            else:
                total_turkce = 0
                total_matematik = 0
                total_fen = 0
                total_sosyal = 0
                total_toplam = 0
                exam_count = 0

                for row in sheet_exam.iter_rows(min_row=2, values_only=True): #first row is header

                    if len(row) >= 6 and row[0] is not None:
                        try:                            
                            total_turkce += (row[1] if row[1] is not None else 0)
                            total_matematik += (row[2] if row[2] is not None else 0)
                            total_fen += (row[3] if row[3] is not None else 0)
                            total_sosyal += (row[4] if row[4] is not None else 0)
                            total_toplam += (row[5] if row[5] is not None else 0)
                            exam_count += 1
                        except TypeError as e:
                            print(f"ERROR: {EXAM_RESULT_EXCEL_PATH} non-numeric value! Row: {row}, error: {e}")
                            continue

                if exam_count > 0:
                    avg_turkce = total_turkce / exam_count
                    avg_matematik = total_matematik / exam_count
                    avg_fen = total_fen / exam_count
                    avg_sosyal = total_sosyal / exam_count
                    avg_toplam = total_toplam / exam_count
                    
                    # calculate success percentages
                    # order is important: Türkçe, Matematik, Fen, Sosyal, Toplam
                    lesson_labels = ["Türkçe", "Matematik", "Fen", "Sosyal", "Toplam"]
                    success_percentages.append(round((avg_turkce / MAX_TURKCE_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_matematik / MAX_MATEMATIK_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_fen / MAX_FEN_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_sosyal / MAX_SOCIAL_QUESTIONS) * 100, 2))
                    success_percentages.append(round((avg_toplam / TOTAL_QUESTIONS) * 100, 2))
                else:
                    (lesson_labels, success_percentages) = GetDefaultLessonDatas()
            
        except Exception as e:
            print(f"ERROR: {e}")
            (lesson_labels, success_percentages) = GetDefaultLessonDatas()
    else:
        print(f"{EXAM_RESULT_EXCEL_PATH} does not exist!")
        (lesson_labels, success_percentages) = GetDefaultLessonDatas()

    # Read weekly schedule excel file
    table_headers = []
    table_rows = []

    if os.path.exists(SCHEDULE_EXCEL_PATH):
        try:
            workbook_table = openpyxl.load_workbook(SCHEDULE_EXCEL_PATH)
            sheet_table = workbook_table.active

            table_headers = [cell.value for cell in sheet_table[1]]

            for row in sheet_table.iter_rows(min_row=2, values_only=True):
                # if value is None, replace with empty string, else convert to string
                cleaned_row = ["" if val is None else str(val) for val in row] 
                table_rows.append(cleaned_row)
            
        except Exception as e:
            print(f"It could not read table datas: {e}")
            table_headers = ["Error"]
            table_rows = [["Invalid Data"]]
    else:
        print(f"Not found excel file: {SCHEDULE_EXCEL_PATH}")
        table_headers = ["Error"]
        table_rows = [["File Not Found"]]

    # Create the main page with the data
    return render_template('index.html', 
                           lesson_labels=lesson_labels,
                           success_percentages=success_percentages,
                           table_headers=table_headers,
                           table_rows=table_rows)

# ---------- CHAT ROUTER ----------
@app.route('/send_chat_message', methods=['POST'])
def send_chat_message():
    user_message = request.json.get('message') # Get message from JSON request

    ai_response = connect_ai.aiAnswer(user_message) # Process the user message with AI
    
    return jsonify({'response': ai_response}) # Return JSON response -> AI response

# ---------- ADD SCORE ROUTER ----------
@app.route('/add_net') 
def add_net_page(): # ---------- ADD SCORE PAGE / ADD_NET_PAGE ----------
    
    deneme_labels = []
    turkce_score = []
    matematik_score = []
    fen_score = []
    sosyal_score = []
    total_score = []

    if os.path.exists(EXAM_RESULT_EXCEL_PATH):
        try:
            workbook_exam = openpyxl.load_workbook(EXAM_RESULT_EXCEL_PATH)
            sheet_exam = workbook_exam.active
            for row in sheet_exam.iter_rows(min_row=2, values_only=True):
                
                if len(row) >= 6: # at least 6 columns
                    
                    if row[0] is None:
                        deneme_labels.append("Deneme İsmi Yok") # exam name is None
                    else:
                        deneme_labels.append(row[0]) # exam name

                    turkce_score.append(row[1] if row[1] is not None else 0)
                    matematik_score.append(row[2] if row[2] is not None else 0)
                    fen_score.append(row[3] if row[3] is not None else 0)
                    sosyal_score.append(row[4] if row[4] is not None else 0)
                    total_score.append(row[5] if row[5] is not None else 0)
            
        except Exception as e:
            print(f"Error: {e}")
            (deneme_labels, turkce_score, matematik_score, fen_score, sosyal_score, total_score) = EmptyExamResultRow()
            
    else:
        print(f"Çizgi Grafik (deneme_neti.xlsx) Excel dosyası bulunamadı: {EXAM_RESULT_EXCEL_PATH}")
        (deneme_labels, turkce_score, matematik_score, fen_score, sosyal_score, total_score) = EmptyExamResultRow()

    # Create the add net page with the data
    return render_template('add_net.html',
                           deneme_labels=deneme_labels,
                           turkce_score=turkce_score,
                           matematik_score=matematik_score,
                           fen_score=fen_score,
                           sosyal_score=sosyal_score,
                           total_score=total_score,
                           MAX_TURKCE_NET=MAX_TURKCE_QUESTIONS,
                           MAX_MATEMATIK_NET=MAX_MATEMATIK_QUESTIONS,
                           MAX_FEN_NET=MAX_FEN_QUESTIONS,
                           MAX_SOCIAL_NET=MAX_SOCIAL_QUESTIONS,
                           MAX_TOTAL_NET=TOTAL_QUESTIONS)

# ---------- PROCESS ADD NET ROUTER ----------
@app.route('/process_add_net', methods=['POST'])
def process_add_net(): # Reload the add net page after processing the form data -> to rewrite excel

    if request.method == 'POST':
        exam_name = request.form.get('exam_name').strip()

        # These datas are string, convert them to int
        turkce_score = request.form.get('turkce_score')
        matematik_score = request.form.get('matematik_score')
        fen_score = request.form.get('fen_score')
        sosyal_score = request.form.get('sosyal_score')

        if not exam_name:
            flash('Lütfen deneme adını doldurun!', 'error')
            return redirect(url_for('add_net_page'))

        try:
            # Convert scores to integers, if empty set to 0
            turkce_score = int(turkce_score) if turkce_score else 0
            matematik_score = int(matematik_score) if matematik_score else 0
            fen_score = int(fen_score) if fen_score else 0
            sosyal_score = int(sosyal_score) if sosyal_score else 0
            
            # input range checks
            if not (MAX_TURKCE_QUESTIONS * -0.25 <= turkce_score <= MAX_TURKCE_QUESTIONS):
                flash(f'Maksimum Türkçe neti: {MAX_TURKCE_QUESTIONS}!', 'error')
                return redirect(url_for('add_net_page'))
            if not (MAX_MATEMATIK_QUESTIONS * -0.25 <= matematik_score <= MAX_MATEMATIK_QUESTIONS):
                flash(f'Maksimum matematil neti: {MAX_MATEMATIK_QUESTIONS}!', 'error')
                return redirect(url_for('add_net_page'))
            if not (MAX_FEN_QUESTIONS * -0.25 <= fen_score <= MAX_FEN_QUESTIONS):
                flash(f'Maksimum fen neti: {MAX_FEN_QUESTIONS}!', 'error')
                return redirect(url_for('add_net_page'))
            if not (MAX_SOCIAL_QUESTIONS * -0.25 <= sosyal_score <= MAX_SOCIAL_QUESTIONS):
                flash(f'Maksimum sosyal neti: {MAX_SOCIAL_QUESTIONS}!', 'error')
                return redirect(url_for('add_net_page'))
            
            total_score = turkce_score + matematik_score + fen_score + sosyal_score

            # If excel file does not exist, create it
            if not os.path.exists(EXAM_RESULT_EXCEL_PATH):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Ders Netleri"
                sheet.append(["Deneme İsmi", "Türkçe Neti", "Matematik Neti", "Fen Neti", "Sosyal Neti", "Toplam Net"])
                workbook.save(EXAM_RESULT_EXCEL_PATH)
            else:
                workbook = openpyxl.load_workbook(EXAM_RESULT_EXCEL_PATH)
                sheet = workbook.active

            # check if the exam name already exists
            found_row_idx = -1 # -1 means not found -> it is a FLAG
            for r_idx in range(2, sheet.max_row + 1): # 2. satırdan başla
                if sheet.cell(row=r_idx, column=1).value == exam_name: # find exam name
                    found_row_idx = r_idx
                    break
            
            if found_row_idx != -1: # exam name found
                sheet.cell(row=found_row_idx, column=2).value = turkce_score
                sheet.cell(row=found_row_idx, column=3).value = matematik_score
                sheet.cell(row=found_row_idx, column=4).value = fen_score
                sheet.cell(row=found_row_idx, column=5).value = sosyal_score
                sheet.cell(row=found_row_idx, column=6).value = total_score
                flash(f'{exam_name} sınav dosyası güncellendi!', 'success')
            else: # exam name not found, append a new row
                new_row_data = [exam_name, turkce_score, matematik_score, fen_score, sosyal_score, total_score]
                sheet.append(new_row_data) # on last row
                flash(f'{exam_name} denemesi eklendi!', 'success')
            
            workbook.save(EXAM_RESULT_EXCEL_PATH)
            
            return redirect(url_for('add_net_page')) # return to add net page
        
        except ValueError:
            flash('Net değerleri geçerli sayılar olmalıdır.', 'error')
            return redirect(url_for('add_net_page'))
        except Exception as e:
            flash(f'Veri kaydedilirken bir hata oluştu: {e}', 'error')
            print(f"Writing error: {e}")
            return redirect(url_for('add_net_page'))
    
    return redirect(url_for('add_net_page'))

# ---------- Run the Flask application ----------
if __name__ == '__main__':
    app.run(debug=True)